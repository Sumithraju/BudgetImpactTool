"""Reading an uploaded CSV or workbook into a table, once — M19 section 5.1.

Extracted when a second importer needed the same behaviour. Everything here is
about getting bytes into rows and reporting where a problem was; what the rows
*mean* belongs to the importer that asked for them.

Two properties the importers depend on and that live here rather than in each
of them: a cell reference on every finding, because a message without one is
not actionable, and `data_only` reading, because a workbook saved without
cached formula values reads as empty and guessing around that is how a model
silently loses half its inputs.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Final

from ..constants.workbook import (
    CSV_SHEET_NAME,
    MAX_UPLOAD_BYTES,
    FindingCode,
    FindingSeverity,
)
from ..schemas.comparator_import import CellRef, ImportFinding

_XLSX_MAGIC: Final[bytes] = b"PK\x03\x04"
HEADER_ROW: Final[int] = 1


def column_letter(index: int) -> str:
    """0-based column index to a spreadsheet letter. 0 -> A, 26 -> AA."""
    letters = ""
    n = index + 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


@dataclass
class FindingCollector:
    """Accumulates findings with their cell references.

    Findings accumulate rather than raise so an analyst fixing a fifty-row
    sheet gets fifty messages in one pass; a parser that stops at the first
    error turns one correction into fifty round trips (M19 section 5.4).
    """

    sheet: str = CSV_SHEET_NAME
    findings: list[ImportFinding] = field(default_factory=list)

    def add(
        self,
        severity: FindingSeverity,
        code: FindingCode,
        message: str,
        *,
        ref: CellRef | None = None,
        supplied: str | None = None,
        expected: str | None = None,
    ) -> None:
        self.findings.append(
            ImportFinding(
                severity=severity, code=code, message=message,
                ref=ref, supplied=supplied, expected=expected,
            )
        )

    def cell(self, column_index: int, row_number: int, label: str | None = None) -> CellRef:
        return CellRef(
            sheet=self.sheet,
            cell=f"{column_letter(column_index)}{row_number}",
            column_label=label,
            row_number=row_number,
        )

    @property
    def has_error(self) -> bool:
        return any(f.severity is FindingSeverity.ERROR for f in self.findings)


@dataclass(frozen=True)
class WorkbookTable:
    sheet: str
    header: list[str]
    rows: list[list[Any]]


def read_table(data: bytes, collector: FindingCollector) -> WorkbookTable | None:
    """Bytes to a header and rows, or `None` with the reason recorded.

    The file kind is decided by content, not by the filename's extension: a
    .xlsx renamed .csv is still a workbook, and trusting the extension turns
    that into an unreadable-encoding error rather than a useful one.
    """
    if len(data) > MAX_UPLOAD_BYTES:
        collector.add(
            FindingSeverity.ERROR, FindingCode.FILE_TOO_LARGE,
            "That file is larger than this import accepts.",
            supplied=f"{len(data):,} bytes",
            expected=f"at most {MAX_UPLOAD_BYTES:,} bytes",
        )
        return None

    if not data:
        collector.add(FindingSeverity.ERROR, FindingCode.EMPTY_FILE, "That file is empty.")
        return None

    return (
        _read_xlsx(data, collector)
        if data.startswith(_XLSX_MAGIC)
        else _read_csv(data, collector)
    )


def _read_csv(data: bytes, collector: FindingCollector) -> WorkbookTable | None:
    import csv
    import io

    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        collector.add(
            FindingSeverity.ERROR, FindingCode.NOT_A_WORKBOOK,
            "That file is neither a readable CSV nor an .xlsx workbook.",
        )
        return None

    reader = csv.reader(io.StringIO(text))
    table = [row for row in reader if any(cell.strip() for cell in row)]
    if not table:
        collector.add(FindingSeverity.ERROR, FindingCode.EMPTY_FILE, "That file is empty.")
        return None

    return WorkbookTable(
        sheet=CSV_SHEET_NAME,
        header=[c.strip() for c in table[0]],
        rows=[list(r) for r in table[1:]],
    )


def _read_xlsx(data: bytes, collector: FindingCollector) -> WorkbookTable | None:
    import io

    from openpyxl import load_workbook

    try:
        # data_only: values, not formulas.
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:                                  # noqa: BLE001
        collector.add(
            FindingSeverity.ERROR, FindingCode.NOT_A_WORKBOOK,
            "That file could not be opened as a workbook. If it is password "
            "protected, remove the protection and try again.",
        )
        return None

    sheet = workbook.worksheets[0]
    collector.sheet = sheet.title
    table = [
        list(row)
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None and str(cell).strip() for cell in row)
    ]
    workbook.close()

    if not table:
        collector.add(FindingSeverity.ERROR, FindingCode.EMPTY_FILE, "That sheet is empty.")
        return None

    body = table[1:]
    if body and all(all(c is None for c in row) for row in body):
        collector.add(
            FindingSeverity.ERROR, FindingCode.NO_CACHED_VALUE,
            "Every value in this sheet is a formula with no cached result. "
            "Open it in Excel, save it, and upload it again.",
        )
        return None

    return WorkbookTable(
        sheet=sheet.title,
        header=[str(c).strip() if c is not None else "" for c in table[0]],
        rows=body,
    )
