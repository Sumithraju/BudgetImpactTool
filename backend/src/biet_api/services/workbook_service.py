# mypy: disable-error-code="no-untyped-call,import-untyped,no-untyped-def,attr-defined,assignment,index,arg-type,union-attr"
#
# openpyxl ships no type stubs. Scoped to this module, as in `excel_service`;
# the rest of the package stays under full --strict.
"""Workbook import and templates — M19.

HEOR runs on Excel. A tool that requires re-typing a market model somebody has
already built does not fit the workflow it is meant to serve, so this module
accepts one.

Three decisions shape it:

**The template is generated from the field dictionary, not written by hand.**
Every column carries the same description that appears as hover text in the
interface, because they are literally the same string. A template that drifts
from the tool is worse than no template — the analyst fills in a column that no
longer exists and discovers it after the import.

**A bad row is reported, not fatal.** A file with three good sheets and one bad
row is mostly right, and rejecting it wholesale wastes the ninety-seven percent
that parsed. Every rejected cell comes back with its sheet, row and column, so
the person fixing it is given an instruction rather than a puzzle.

**Nothing is saved by importing.** The import returns a parsed *draft*. An
import that silently created a scenario would give nobody a chance to see what
the file was read as, and every imported value carries its sheet and row as
provenance precisely so that review is possible.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..constants.domain import Perspective
from ..constants.field_guide import FIELD_GROUPS, FieldSpec
from ..constants.parameter_paths import spec_for
from ..schemas.pricing import ImportedScenario, ImportIssue, ImportResponse, PriceEdit

ACCENT = "07707C"
INK = "0E181B"
INK_2 = "42585F"
SURFACE_2 = "EDF3F4"
LINE = "D8E3E6"
INPUT_FILL = "FFF9E8"

_thin = Side(style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

H1 = Font(name="Calibri", size=16, bold=True, color=INK)
H2 = Font(name="Calibri", size=11, bold=True, color=ACCENT)
TH = Font(name="Calibri", size=9, bold=True, color=INK_2)
TD = Font(name="Calibri", size=10, color=INK)
SMALL = Font(name="Calibri", size=8.5, color=INK_2)

FILL_HEAD = PatternFill("solid", fgColor=SURFACE_2)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_FILL)
WRAP = Alignment(wrap_text=True, vertical="top")

SHEET_SCENARIO = "Scenario"
SHEET_SUBGROUPS = "Subgroups"

#: `subgroup_1_diabesity_cases` -> `diabesity`. Matched by pattern rather than
#: listed, because the analyst's file numbers its subgroups while this model
#: names them, and a fixed column list would break the first time one moved.
#: Capturing the *index* as well as the slug matters: the cases column names
#: its subgroup (`subgroup_1_diabesity_cases`) while the eligibility column
#: beside it does not (`subgroup_1_clinically_eligible`). The pair is found by
#: index, so deriving one column name from the other by string surgery — which
#: silently produces a column that does not exist — is not needed.
_SUBGROUP_CASES = re.compile(r"subgroup_(\d+)_(.+)_cases")

#: The analyst's column slugs mapped onto this model's subgroup codes. Explicit
#: rather than fuzzy-matched: silently attaching an unrecognised column to the
#: nearest-looking subgroup would import the wrong population.
_SUBGROUP_ALIASES: dict[str, str] = {
    "diabesity": "diabesity",
    "cvd_mace_risk": "cvd_mace_risk",
    "htn_osa": "htn_osa",
    "severe_obesity_class2_3": "severe_obesity",
    "severe_obesity": "severe_obesity",
}
SHEET_PRICES = "Prices"
SHEET_GUIDE = "Guide"

#: Column headers on the Scenario sheet. `group` and `guidance` are there for
#: the person filling it in; only `field` and `value` are read back.
SCENARIO_COLUMNS = ("Section", "Input", "Value", "Unit", "Example", "What it means")
PRICE_COLUMNS = (
    "drug_id", "Therapy", "Market", "Currency",
    "Unit price", "Annual cost", "Basis today", "Note",
)


class WorkbookService:
    """Templates out, drafts in."""

    # ----------------------------------------------------------------- out

    def scenario_template(self, prices: list[Any] | None = None) -> bytes:
        """A workbook an analyst can fill in and send back.

        The price sheet is pre-filled with the model's current grid rather than
        left blank. An empty sheet asks the analyst to supply ten markets'
        prices from nothing; a pre-filled one asks them to correct the three
        they disagree with, which is both less work and a better question.
        """
        workbook = Workbook()
        self._scenario_sheet(workbook.active)
        self._price_sheet(workbook.create_sheet(SHEET_PRICES), prices or [])
        self._guide_sheet(workbook.create_sheet(SHEET_GUIDE))

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _scenario_sheet(self, sheet) -> None:
        sheet.title = SHEET_SCENARIO
        sheet["A1"] = "BIET — scenario inputs"
        sheet["A1"].font = H1
        sheet["A2"] = (
            "Fill in the Value column. Leave a row blank to use the model's "
            "seeded default — a blank and a value that happens to equal the "
            "default are different claims, and the model records them "
            "differently. Rates are fractions: 0.19 means 19%."
        )
        sheet["A2"].font = SMALL
        sheet["A2"].alignment = WRAP
        sheet.merge_cells("A2:F2")
        sheet.row_dimensions[2].height = 42

        widths = (18, 26, 18, 20, 24, 74)
        for index, (label, width) in enumerate(zip(SCENARIO_COLUMNS, widths, strict=True), 1):
            cell = sheet.cell(row=4, column=index, value=label)
            cell.font = TH
            cell.fill = FILL_HEAD
            cell.border = BORDER
            sheet.column_dimensions[get_column_letter(index)].width = width

        row = 5
        for group in FIELD_GROUPS:
            for spec in group.fields:
                if not spec.importable:
                    continue
                sheet.cell(row=row, column=1, value=group.label).font = TD
                sheet.cell(row=row, column=2, value=spec.label).font = TD
                value_cell = sheet.cell(row=row, column=3)
                value_cell.fill = FILL_INPUT
                value_cell.border = BORDER
                # The key the importer reads back, kept out of sight rather
                # than out of the file: a template whose columns can be
                # reordered by the person filling it in cannot be parsed
                # positionally, and asking them not to reorder is not a plan.
                sheet.cell(row=row, column=8, value=spec.key).font = SMALL
                sheet.cell(row=row, column=4, value=spec.unit or "—").font = SMALL
                sheet.cell(row=row, column=5, value=spec.example or "").font = SMALL
                guidance = spec.description + (
                    f"  {spec.effect}" if spec.effect else ""
                )
                note = sheet.cell(row=row, column=6, value=guidance)
                note.font = SMALL
                note.alignment = WRAP
                sheet.row_dimensions[row].height = 30
                row += 1

        sheet.column_dimensions["H"].hidden = True
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(p.value for p in Perspective) + '"',
            allow_blank=True,
        )
        sheet.add_data_validation(validation)
        for candidate in range(5, row):
            if sheet.cell(row=candidate, column=8).value == "perspective":
                validation.add(sheet.cell(row=candidate, column=3))

        sheet.freeze_panes = "A5"

    def _price_sheet(self, sheet, prices: list[Any]) -> None:
        sheet["A1"] = "Prices — one row per therapy per market"
        sheet["A1"].font = H1
        sheet["A2"] = (
            "Edit either the unit price or the annual cost; the other is "
            "derived from the therapy's regimen. Rows marked 'derived' have no "
            "observed price in that market and are the ones worth your "
            "attention first. An edit becomes a scenario override, not a "
            "change to the reference data."
        )
        sheet["A2"].font = SMALL
        sheet["A2"].alignment = WRAP
        sheet.merge_cells("A2:H2")
        sheet.row_dimensions[2].height = 42

        widths = (10, 30, 10, 10, 16, 16, 16, 44)
        for index, (label, width) in enumerate(zip(PRICE_COLUMNS, widths, strict=True), 1):
            cell = sheet.cell(row=4, column=index, value=label)
            cell.font = TH
            cell.fill = FILL_HEAD
            cell.border = BORDER
            sheet.column_dimensions[get_column_letter(index)].width = width

        for offset, price in enumerate(prices):
            row = 5 + offset
            sheet.cell(row=row, column=1, value=price.drug_id).font = SMALL
            sheet.cell(row=row, column=2, value=price.drug_name).font = TD
            sheet.cell(row=row, column=3, value=price.country_code).font = TD
            sheet.cell(row=row, column=4, value=price.currency_code).font = SMALL
            unit = sheet.cell(row=row, column=5, value=round(price.unit_price, 4))
            unit.fill = FILL_INPUT
            unit.border = BORDER
            # A therapy with no regimen (the "no pharmacotherapy" comparator)
            # has no annual cost to state. Left blank rather than written as
            # zero: a blank asks to be filled in, a zero asserts the therapy
            # is free.
            annual = sheet.cell(
                row=row, column=6,
                value=round(price.annual_cost, 2) if price.annual_cost > 0 else None,
            )
            annual.fill = FILL_INPUT
            annual.border = BORDER
            sheet.cell(
                row=row, column=7,
                value="observed" if price.is_observed else "derived",
            ).font = SMALL
            sheet.cell(row=row, column=8, value=price.source[:200]).font = SMALL

        sheet.freeze_panes = "A5"

    def _guide_sheet(self, sheet) -> None:
        """The taxonomy in full, including the fields no template can carry.

        Outcome evidence has no input column — a relative risk reduction
        belongs to a trial, not to a scenario — but it is still what a reader
        will ask about, so it is documented here rather than omitted.
        """
        sheet["A1"] = "What every input means"
        sheet["A1"].font = H1
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 88

        row = 3
        for group in FIELD_GROUPS:
            header = sheet.cell(row=row, column=1, value=group.label)
            header.font = H2
            summary = sheet.cell(row=row, column=3, value=group.summary)
            summary.font = SMALL
            summary.alignment = WRAP
            row += 1
            for spec in group.fields:
                sheet.cell(row=row, column=2, value=spec.label).font = TD
                text = spec.description
                if spec.effect:
                    text += f"\n\nWhy it matters: {spec.effect}"
                if spec.typical_range:
                    text += f"\n\nTypical range: {spec.typical_range}"
                cell = sheet.cell(row=row, column=3, value=text)
                cell.font = SMALL
                cell.alignment = WRAP
                sheet.row_dimensions[row].height = 58
                row += 1
            row += 1

    def scenario_template_csv(self) -> str:
        """The same template, flattened, for anyone not on Excel."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["field", "value", "section", "input", "unit", "example", "what_it_means"])
        for group in FIELD_GROUPS:
            for spec in group.fields:
                if not spec.importable:
                    continue
                writer.writerow([
                    spec.key, "", group.label, spec.label, spec.unit or "",
                    spec.example or "",
                    spec.description + (f" {spec.effect}" if spec.effect else ""),
                ])
        return buffer.getvalue()

    def price_template_csv(self, prices: list[Any]) -> str:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            "drug_id", "drug_name", "country_code", "currency_code",
            "unit_price", "annual_cost", "basis_today", "source",
        ])
        for price in prices:
            writer.writerow([
                price.drug_id, price.drug_name, price.country_code,
                price.currency_code, round(price.unit_price, 6),
                round(price.annual_cost, 2),
                "observed" if price.is_observed else "derived",
                price.source[:300],
            ])
        return buffer.getvalue()

    # ----------------------------------------------------------------- in

    def parse(self, payload: bytes, filename: str) -> ImportResponse:
        """Read a workbook or CSV into a scenario draft.

        Dispatches on the extension rather than sniffing the bytes, because a
        misidentified file should fail with "that is not a workbook" and not
        with a parser error forty rows in.
        """
        if filename.lower().endswith(".csv"):
            return self._parse_csv(payload)
        return self._parse_xlsx(payload)

    def _parse_csv(self, payload: bytes) -> ImportResponse:
        """A CSV is either this tool's own field/value template or the
        analyst's own subgroup derivation. Which one is decided by the columns
        present, not by asking — a file that has to be renamed before it can be
        read is a file the analyst will not send."""
        text = payload.decode("utf-8-sig", errors="replace")
        table = list(csv.DictReader(io.StringIO(text)))
        draft = ImportedScenario()
        issues: list[ImportIssue] = []

        read = self._parse_subgroup_rows(table, draft, issues)
        if read:
            return ImportResponse(
                scenario=draft, issues=issues, rows_read=read,
                accepted=not any(i.severity == "error" for i in issues),
            )

        rows = 0
        for index, row in enumerate(table, start=2):
            rows += 1
            key = (row.get("field") or "").strip()
            raw = (row.get("value") or "").strip()
            if not key or not raw:
                continue
            self._apply(draft, key, raw, SHEET_SCENARIO, index, issues)

        return ImportResponse(
            scenario=draft, issues=issues, rows_read=rows,
            accepted=not any(i.severity == "error" for i in issues),
        )

    def _parse_xlsx(self, payload: bytes) -> ImportResponse:
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
        draft = ImportedScenario()
        issues: list[ImportIssue] = []
        rows = 0

        if SHEET_SCENARIO in workbook.sheetnames:
            sheet = workbook[SHEET_SCENARIO]
            for row in range(5, sheet.max_row + 1):
                key = sheet.cell(row=row, column=8).value
                value = sheet.cell(row=row, column=3).value
                if not key:
                    continue
                rows += 1
                if value in (None, ""):
                    continue
                self._apply(draft, str(key), value, SHEET_SCENARIO, row, issues)
        elif not any(
            _SUBGROUP_CASES.fullmatch(column)
            for name in workbook.sheetnames
            for column in _header_of(workbook[name])
        ):
            issues.append(ImportIssue(
                sheet=SHEET_SCENARIO, row=0, column="—", value=None,
                message=(
                    f"The file has no sheet named {SHEET_SCENARIO!r}. Download "
                    "the template and fill that in, or rename the sheet to "
                    "match — the importer reads by sheet name so that column "
                    "order does not matter."
                ),
            ))

        # Any sheet carrying the analyst's own subgroup columns is read as
        # one, whatever it is called. The derivation file is usually a single
        # unnamed sheet, and insisting it be renamed "Subgroups" would be
        # insisting on a step that serves nobody.
        for name in workbook.sheetnames:
            table = _sheet_rows(workbook[name])
            read = self._parse_subgroup_rows(table, draft, issues)
            if read:
                rows += read
                break

        if SHEET_PRICES in workbook.sheetnames:
            sheet = workbook[SHEET_PRICES]
            for row in range(5, sheet.max_row + 1):
                drug_id = sheet.cell(row=row, column=1).value
                if drug_id in (None, ""):
                    continue
                rows += 1
                self._apply_price(draft, sheet, row, issues)

        return ImportResponse(
            scenario=draft, issues=issues, rows_read=rows,
            accepted=not any(i.severity == "error" for i in issues),
        )

    # ------------------------------------------------- WHO-derivation shape

    def _parse_subgroup_rows(
        self,
        rows: list[dict[str, str]],
        draft: ImportedScenario,
        issues: list[ImportIssue],
    ) -> int:
        """Read a WHO-derived subgroup file into per-market subgroup inputs.

        This is the shape of the file the analyst actually has:
        `subgroup_1_diabesity_cases`, `subgroup_1_clinically_eligible`, and so
        on, one row per country. It is not this tool's own template, and
        requiring the analyst to transpose it into one would be requiring them
        to do the import by hand first.

        Columns are therefore recognised by pattern. Each subgroup contributes
        two values per market — its share of the prevalent population, and the
        fraction of that share which is clinically eligible — and both arrive as
        editable overrides rather than as silent reference data, because they
        are the analyst's derivation and not a published figure.
        """
        if not rows:
            return 0

        cases = {
            column: (match.group(1), match.group(2))
            for column in rows[0]
            if (match := _SUBGROUP_CASES.fullmatch(column))
        }
        if not cases:
            return 0

        read = 0
        for index, row in enumerate(rows, start=2):
            code = (row.get("iso_code") or row.get("country_code") or "").strip().upper()
            if len(code) != 3:
                continue
            total = _as_float(row.get("total_obese_prevalent_cases"))
            read += 1

            prevalence = _as_float(row.get("adult_obesity_prevalence_pct"))
            if prevalence is not None:
                # The column is a percentage by its own name; this model works
                # in fractions everywhere. Divided once, here.
                draft.overrides.append({
                    "parameter_path": "epidemiology.prevalence",
                    "country_code": code,
                    "value": round(prevalence / 100.0, 6),
                    "note": f"Imported for {code} from the subgroup file, row {index}.",
                })

            for column, (position, slug) in sorted(cases.items()):
                subgroup = _SUBGROUP_ALIASES.get(slug)
                if subgroup is None:
                    issues.append(ImportIssue(
                        sheet=SHEET_SUBGROUPS, row=index, column=column, value=slug,
                        message=(
                            f"{slug!r} does not match a subgroup this model knows. "
                            "Its cases are read but not applied — rename the column "
                            "or add the subgroup before relying on it."
                        ),
                        severity="warning",
                    ))
                    continue

                count = _as_float(row.get(column))
                eligible = _as_float(row.get(f"subgroup_{position}_clinically_eligible"))
                if count is None or not total:
                    continue

                draft.overrides.append({
                    "parameter_path": f"subgroup.{subgroup}.share",
                    "country_code": code,
                    "value": round(count / total, 6),
                    "note": (
                        f"{count:,.0f} of {total:,.0f} prevalent cases in {code}, "
                        f"imported from row {index}."
                    ),
                })
                if eligible is not None and count:
                    draft.overrides.append({
                        "parameter_path": f"subgroup.{subgroup}.eligible_factor",
                        "country_code": code,
                        "value": round(eligible / count, 6),
                        "note": (
                            f"{eligible:,.0f} of {count:,.0f} clinically eligible in "
                            f"{code}, imported from row {index}."
                        ),
                    })
        return read

    # ----------------------------------------------------------------- cells

    def _apply(
        self,
        draft: ImportedScenario,
        key: str,
        raw: object,
        sheet: str,
        row: int,
        issues: list[ImportIssue],
    ) -> None:
        """One cell, validated against the same rules the API enforces.

        Deliberately routed through `spec_for` for anything that maps to an
        override path, so an imported rate is bounded exactly as a typed one
        is. Two validators for one rule is how a workbook ends up able to
        express a scenario the API would reject.
        """
        text = str(raw).strip()
        try:
            match key:
                case "country_codes":
                    codes = [
                        part.strip().upper()
                        for part in text.replace(";", ",").split(",")
                        if part.strip()
                    ]
                    if not codes:
                        raise ValueError("no market codes found")
                    draft.country_codes = codes
                case "subgroup_codes":
                    draft.subgroup_codes = [
                        part.strip()
                        for part in text.replace(";", ",").split(",")
                        if part.strip()
                    ]
                case "asset_name":
                    draft.asset_name = text
                case "launch_year":
                    draft.launch_year = int(float(text))
                case "horizon_years":
                    horizon = int(float(text))
                    if not 1 <= horizon <= 5:
                        raise ValueError("horizon must be between 1 and 5 years")
                    draft.horizon_years = horizon
                case "reporting_currency":
                    if len(text) != 3:
                        raise ValueError("currency must be a three-letter code")
                    draft.reporting_currency = text.upper()
                case "perspective":
                    draft.perspective = Perspective(text.lower()).value
                case "covered_population":
                    population = int(float(text.replace(",", "")))
                    if population <= 0:
                        raise ValueError("covered lives must be greater than zero")
                    draft.covered_population = population
                case _:
                    self._apply_override(draft, key, text, sheet, row, issues)
        except (ValueError, TypeError) as error:
            issues.append(ImportIssue(
                sheet=sheet, row=row, column=key, value=text,
                message=f"{text!r} is not a valid {key.replace('_', ' ')} — {error}",
            ))

    def _apply_override(
        self,
        draft: ImportedScenario,
        key: str,
        text: str,
        sheet: str,
        row: int,
        issues: list[ImportIssue],
    ) -> None:
        from ..constants.field_guide import FIELD_INDEX

        field: FieldSpec | None = FIELD_INDEX.get(key)
        if field is None or field.parameter_path is None:
            issues.append(ImportIssue(
                sheet=sheet, row=row, column=key, value=text,
                message=(
                    f"{key!r} is not an input this model accepts. Download a "
                    "fresh template — the field list has probably moved on "
                    "since this file was made."
                ),
                severity="warning",
            ))
            return

        # Templated paths (`criteria.<code>.factor`) are per-item and have no
        # single cell in a flat sheet. Named as unsupported rather than
        # silently skipped: an override that quietly does nothing is worse than
        # one that errors, because the analyst believes they changed an
        # assumption when they did not.
        if "<" in field.parameter_path:
            issues.append(ImportIssue(
                sheet=sheet, row=row, column=key, value=text,
                message=(
                    f"{field.label} varies per therapy or per criterion, so it "
                    "cannot be set from a single cell. Use the Prices sheet, or "
                    "set it in the interface."
                ),
                severity="warning",
            ))
            return

        value = float(text.replace("%", "").replace(",", ""))
        if "%" in text:
            value /= 100.0

        spec = spec_for(field.parameter_path)
        if spec is not None and spec.minimum is not None:
            below = value <= spec.minimum if spec.exclusive_min else value < spec.minimum
            above = (
                spec.maximum is not None
                and (value >= spec.maximum if spec.exclusive_max else value > spec.maximum)
            )
            if below or above:
                bounds = f"{spec.minimum} to {spec.maximum}"
                raise ValueError(
                    f"outside the accepted range ({bounds}). Rates are "
                    f"fractions here — 19% is 0.19, not 19"
                )

        draft.overrides.append({
            "parameter_path": field.parameter_path,
            "value": value,
            "note": f"Imported from {sheet}, row {row}.",
        })

    def _apply_price(
        self, draft: ImportedScenario, sheet, row: int, issues: list[ImportIssue],
    ) -> None:
        drug_id = sheet.cell(row=row, column=1).value
        country = sheet.cell(row=row, column=3).value
        unit = sheet.cell(row=row, column=5).value
        annual = sheet.cell(row=row, column=6).value

        try:
            # Blank and zero are both "not supplied" here. A zero price is not
            # a price, and a therapy legitimately carrying no annual cost —
            # the "no pharmacotherapy" comparator — would otherwise be
            # rejected as invalid on every single import.
            unit_price = self._optional_positive(unit)
            annual_cost = self._optional_positive(annual)
            code = str(country or "").strip().upper()
            if len(code) != 3:
                raise ValueError(
                    f"{code or 'blank'!r} is not a three-letter market code"
                )
            drug = int(drug_id)
        except (TypeError, ValueError) as error:
            issues.append(ImportIssue(
                sheet=SHEET_PRICES, row=row, column="unit_price",
                value=None if unit is None else str(unit),
                message=f"could not be read as a price — {error}",
            ))
            return

        if unit_price is None and annual_cost is None:
            return
        draft.prices.append(PriceEdit(
            drug_id=drug,
            country_code=code,
            unit_price=unit_price,
            annual_cost=annual_cost,
            note=f"Imported from {SHEET_PRICES}, row {row}.",
        ))

    @staticmethod
    def _optional_positive(raw: object) -> float | None:
        """A cell as a positive number, or None when it is blank or zero.

        Raises on text that is not a number at all, which is a real mistake
        worth reporting — as distinct from an empty cell, which is the ordinary
        way of saying "leave this alone".
        """
        if raw in (None, ""):
            return None
        value = float(str(raw).replace(",", ""))
        return value if value > 0 else None


def _as_float(raw: object) -> float | None:
    """A cell as a number, or None when it is blank or not one.

    Returns None rather than raising: a subgroup file legitimately has gaps — a
    country with no published figure for one comorbidity — and a gap is not an
    error worth stopping an import for.
    """
    if raw in (None, ""):
        return None
    try:
        return float(str(raw).replace(",", "").replace("%", "").strip())
    except ValueError:
        return None


def _header_of(sheet: Any) -> list[str]:
    """The first row that looks like a header, as lowercase column names."""
    for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
        values = [str(v).strip() for v in row if v not in (None, "")]
        if len(values) >= 3:
            return [v.lower() for v in values]
    return []


def _sheet_rows(sheet: Any) -> list[dict[str, str]]:
    """A worksheet as dicts keyed by its header row.

    The header is found rather than assumed to be row 1: the analyst's files
    often carry a title line above it, and failing on that would reject a file
    for a cosmetic reason.
    """
    header: list[str] | None = None
    out: list[dict[str, str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = ["" if v is None else str(v).strip() for v in row]
        if header is None:
            named = [v for v in values if v]
            if len(named) >= 3 and any(
                _SUBGROUP_CASES.fullmatch(v.lower()) for v in named
            ):
                header = [v.lower() for v in values]
            continue
        if not any(values):
            continue
        out.append(dict(zip(header, values, strict=False)))
    return out
