"""Comparator workbook and CSV import — M19 sections 5.1 to 5.6.

Three rules from the module spec shape everything here.

**All findings in one pass.** An analyst fixing a fifty-row sheet should get
fifty messages once; a parser that stops at the first error turns one
correction into fifty round trips (M19 section 5.4).

**A file is accepted whole or not at all.** Any error-severity finding rejects
it and nothing is written. Warnings travel onto the result beside the values
they concern.

**Units are declared, never inferred.** The header says `(%)`, so the boundary
divides by 100 exactly once. A value that cannot be read under its declared
unit is a finding, never a coercion: a prevalence of 20.64 silently read as
2064% and one of 0.2064 read as 0.2% are both catastrophic and both look
plausible in a log (M19 section 5.2).
"""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from typing import Any, Final

from ..constants.workbook import (
    CSV_SHEET_NAME,
    DEFAULT_IMPORT_TIER,
    MAX_ROWS,
    MAX_UPLOAD_BYTES,
    OPTIONAL_COST_COLUMNS,
    PERCENT_DIVISOR,
    REQUIRED_COLUMNS,
    SHARE_TOLERANCE,
    VALID_TIERS,
    FindingCode,
    FindingSeverity,
    WorkbookColumn,
)
from ..schemas.comparator_import import (
    CellRef,
    ComparatorImportResult,
    ImportedComparator,
    ImportFinding,
)

#: A share above this in a `(%)` column cannot be a percentage.
MAX_PERCENT: Final[float] = 100.0
_XLSX_MAGIC: Final[bytes] = b"PK\x03\x04"
_HEADER_ROW: Final[int] = 1


def _column_letter(index: int) -> str:
    """0-based column index to a spreadsheet letter. 0 -> A, 26 -> AA."""
    letters = ""
    n = index + 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


class ComparatorImportService:
    """Parses one uploaded file into validated comparator rows.

    Holds no session. Import is a pure transformation from bytes to a result;
    what happens to an accepted row afterwards is the caller's decision, and
    deliberately not this service's — M19 section 5.6 keeps registry writes
    on M12's promotion path so the registry stays the single record of what a
    drug is.
    """

    def __init__(self, known_markets: frozenset[str], known_currencies: frozenset[str]) -> None:
        self._markets = known_markets
        self._currencies = known_currencies
        self._findings: list[ImportFinding] = []
        self._sheet = CSV_SHEET_NAME

    # ------------------------------------------------------------- findings

    def _add(
        self,
        severity: FindingSeverity,
        code: FindingCode,
        message: str,
        *,
        ref: CellRef | None = None,
        supplied: str | None = None,
        expected: str | None = None,
    ) -> None:
        self._findings.append(
            ImportFinding(
                severity=severity,
                code=code,
                message=message,
                ref=ref,
                supplied=supplied,
                expected=expected,
            )
        )

    def _cell(self, column_index: int, row_number: int, label: str | None = None) -> CellRef:
        return CellRef(
            sheet=self._sheet,
            cell=f"{_column_letter(column_index)}{row_number}",
            column_label=label,
            row_number=row_number,
        )

    @property
    def _has_error(self) -> bool:
        return any(f.severity is FindingSeverity.ERROR for f in self._findings)

    # ---------------------------------------------------------------- parse

    def parse(self, data: bytes, filename: str) -> ComparatorImportResult:
        self._findings = []
        self._sheet = CSV_SHEET_NAME

        if len(data) > MAX_UPLOAD_BYTES:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.FILE_TOO_LARGE,
                "That file is larger than this import accepts.",
                supplied=f"{len(data):,} bytes",
                expected=f"at most {MAX_UPLOAD_BYTES:,} bytes",
            )
            return self._reject(filename)

        if not data:
            self._add(
                FindingSeverity.ERROR, FindingCode.EMPTY_FILE, "That file is empty."
            )
            return self._reject(filename)

        table = (
            self._read_xlsx(data)
            if data.startswith(_XLSX_MAGIC)
            else self._read_csv(data)
        )
        if table is None:
            return self._reject(filename)

        header, rows = table
        if not rows:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.NO_ROWS,
                "The file has a header row but no comparators beneath it.",
            )
            return self._reject(filename)

        index = self._map_columns(header)
        if self._has_error:
            return self._reject(filename)

        comparators = self._read_rows(rows, index)
        totals = self._check_shares(comparators)

        if self._has_error:
            return self._reject(filename, rows_read=len(rows), share_totals=totals)

        return ComparatorImportResult(
            accepted=True,
            filename=filename,
            sheet=self._sheet,
            rows_read=len(rows),
            findings=tuple(self._findings),
            comparators=tuple(comparators),
            share_totals=totals,
        )

    def _reject(
        self,
        filename: str,
        *,
        rows_read: int = 0,
        share_totals: dict[str, float] | None = None,
    ) -> ComparatorImportResult:
        return ComparatorImportResult(
            accepted=False,
            filename=filename,
            sheet=self._sheet,
            rows_read=rows_read,
            findings=tuple(self._findings),
            comparators=(),
            share_totals=share_totals or {},
        )

    # ----------------------------------------------------------- file kinds

    def _read_csv(self, data: bytes) -> tuple[list[str], list[list[str]]] | None:
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.NOT_A_WORKBOOK,
                "That file is neither a readable CSV nor an .xlsx workbook.",
            )
            return None

        reader = csv.reader(io.StringIO(text))
        table = [row for row in reader if any(cell.strip() for cell in row)]
        if not table:
            self._add(FindingSeverity.ERROR, FindingCode.EMPTY_FILE, "That file is empty.")
            return None
        return [c.strip() for c in table[0]], [list(r) for r in table[1:]]

    def _read_xlsx(self, data: bytes) -> tuple[list[str], list[list[Any]]] | None:
        from openpyxl import load_workbook

        try:
            # data_only: values, not formulas. A workbook saved without cached
            # values reads as empty, which section 5.2 refuses to guess around.
            workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception:                                  # noqa: BLE001
            self._add(
                FindingSeverity.ERROR,
                FindingCode.NOT_A_WORKBOOK,
                "That file could not be opened as a workbook. If it is password "
                "protected, remove the protection and try again.",
            )
            return None

        sheet = workbook.worksheets[0]
        self._sheet = sheet.title
        table = [
            list(row)
            for row in sheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
        workbook.close()

        if not table:
            self._add(FindingSeverity.ERROR, FindingCode.EMPTY_FILE, "That sheet is empty.")
            return None

        header = [str(c).strip() if c is not None else "" for c in table[0]]
        body = table[1:]
        if body and all(all(c is None for c in row) for row in body):
            self._add(
                FindingSeverity.ERROR,
                FindingCode.NO_CACHED_VALUE,
                "Every value in this sheet is a formula with no cached result. "
                "Open it in Excel, save it, and upload it again.",
            )
            return None
        return header, body

    # -------------------------------------------------------------- columns

    def _map_columns(self, header: list[str]) -> dict[WorkbookColumn, int]:
        known = {c.value.casefold(): c for c in WorkbookColumn}
        index: dict[WorkbookColumn, int] = {}

        for position, label in enumerate(header):
            column = known.get(label.casefold())
            if column is None:
                if label:
                    self._add(
                        FindingSeverity.WARNING,
                        FindingCode.UNRECOGNISED_COLUMN,
                        f"Column {label!r} is not part of the comparator contract "
                        f"and was ignored.",
                        ref=self._cell(position, _HEADER_ROW, label),
                    )
                continue
            index[column] = position

        for required in REQUIRED_COLUMNS:
            if required not in index:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.MISSING_COLUMN,
                    f"The sheet has no {required.value!r} column.",
                    expected=required.value,
                )
        return index

    # ----------------------------------------------------------------- rows

    def _read_rows(
        self, rows: list[list[Any]], index: dict[WorkbookColumn, int]
    ) -> list[ImportedComparator]:
        out: list[ImportedComparator] = []
        seen: dict[tuple[str, str], int] = {}
        currency_by_market: dict[str, str] = {}

        for offset, row in enumerate(rows[:MAX_ROWS]):
            row_number = offset + _HEADER_ROW + 1
            name = self._text(row, index, WorkbookColumn.NAME, row_number)
            market = self._text(row, index, WorkbookColumn.MARKET, row_number)
            currency = self._text(row, index, WorkbookColumn.CURRENCY, row_number)
            if not (name and market and currency):
                continue

            market = market.upper()
            currency = currency.upper()

            if market not in self._markets:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.UNKNOWN_MARKET,
                    f"{market!r} is not a market this tool covers.",
                    ref=self._cell(index[WorkbookColumn.MARKET], row_number,
                                   WorkbookColumn.MARKET.value),
                    supplied=market,
                    expected=", ".join(sorted(self._markets)),
                )
            if currency not in self._currencies:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.UNKNOWN_CURRENCY,
                    f"{currency!r} is not a currency with a rate in this run.",
                    ref=self._cell(index[WorkbookColumn.CURRENCY], row_number,
                                   WorkbookColumn.CURRENCY.value),
                    supplied=currency,
                )

            # One market calculates in one currency (M20 section 5.6). Two
            # currencies for one market is a mistake, not a conversion job.
            established = currency_by_market.setdefault(market, currency)
            if established != currency:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.MIXED_CURRENCY_COLUMN,
                    f"{market} is priced in both {established} and {currency}. "
                    f"A market calculates in one currency.",
                    ref=self._cell(index[WorkbookColumn.CURRENCY], row_number,
                                   WorkbookColumn.CURRENCY.value),
                )

            key = (name.casefold(), market)
            if key in seen:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.DUPLICATE_ROW,
                    f"{name!r} appears twice for {market} — rows "
                    f"{seen[key]} and {row_number}.",
                    ref=self._cell(index[WorkbookColumn.NAME], row_number,
                                   WorkbookColumn.NAME.value),
                )
            else:
                seen[key] = row_number

            share = self._share(row, index, row_number)
            drug = self._cost(row, index, WorkbookColumn.DRUG_COST, row_number)
            admin, monitoring, ae = (
                self._cost(row, index, column, row_number)
                for column in OPTIONAL_COST_COLUMNS
            )

            if share is None or drug is None:
                continue

            out.append(
                ImportedComparator(
                    name=name,
                    therapy_type=self._text(row, index, WorkbookColumn.TYPE, row_number)
                    or None,
                    country_code=market,
                    currency_code=currency,
                    market_share=share,
                    drug_cost=drug,
                    admin_cost=admin or 0.0,
                    monitoring_cost=monitoring or 0.0,
                    ae_cost=ae or 0.0,
                    total_cost=drug + (admin or 0.0) + (monitoring or 0.0) + (ae or 0.0),
                    source=self._source(row, index, row_number),
                    confidence_tier=self._tier(row, index, row_number),
                    origin=f"{self._sheet}!"
                           f"{_column_letter(index[WorkbookColumn.NAME])}{row_number}",
                )
            )

        if len(rows) > MAX_ROWS:
            self._add(
                FindingSeverity.WARNING,
                FindingCode.NO_ROWS,
                f"Only the first {MAX_ROWS} rows were read.",
                supplied=f"{len(rows)} rows",
            )
        return out

    # ------------------------------------------------------------ cell kinds

    def _raw(
        self, row: list[Any], index: dict[WorkbookColumn, int], column: WorkbookColumn
    ) -> Any:
        position = index.get(column)
        if position is None or position >= len(row):
            return None
        return row[position]

    def _text(
        self,
        row: list[Any],
        index: dict[WorkbookColumn, int],
        column: WorkbookColumn,
        row_number: int,
    ) -> str:
        value = self._raw(row, index, column)
        text = "" if value is None else str(value).strip()
        if not text and column in REQUIRED_COLUMNS:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.MISSING_VALUE,
                f"{column.value} is blank.",
                ref=self._cell(index.get(column, 0), row_number, column.value),
            )
        return text

    def _is_blank(
        self, row: list[Any], index: dict[WorkbookColumn, int], column: WorkbookColumn
    ) -> bool:
        value = self._raw(row, index, column)
        return value is None or str(value).strip() == ""

    def _number(
        self,
        row: list[Any],
        index: dict[WorkbookColumn, int],
        column: WorkbookColumn,
        row_number: int,
    ) -> float | None:
        """`None` means blank *or* unreadable.

        An unreadable cell has already raised `NOT_A_NUMBER`, so callers test
        `_is_blank` before adding a second finding for the same cell — one
        problem, one message.
        """
        value = self._raw(row, index, column)
        if self._is_blank(row, index, column):
            return None
        try:
            # A thousands separator is how Excel stores a number as text; the
            # comma is stripped, but nothing else is coerced.
            return float(str(value).replace(",", "").strip())
        except ValueError:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.NOT_A_NUMBER,
                f"{column.value} is not a number.",
                ref=self._cell(index.get(column, 0), row_number, column.value),
                supplied=str(value),
            )
            return None

    def _share(
        self, row: list[Any], index: dict[WorkbookColumn, int], row_number: int
    ) -> float | None:
        column = WorkbookColumn.MARKET_SHARE_PCT
        percent = self._number(row, index, column, row_number)
        if percent is None:
            if self._is_blank(row, index, column):
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.MISSING_VALUE,
                    f"{column.value} is blank.",
                    ref=self._cell(index.get(column, 0), row_number, column.value),
                )
            return None
        if percent < 0 or percent > MAX_PERCENT:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.AMBIGUOUS_RATE_UNIT,
                f"{percent:g} cannot be a percentage. This column is labelled "
                f"'{column.value}', so 40 means 40% and 0.4 means 0.4%.",
                ref=self._cell(index.get(column, 0), row_number, column.value),
                supplied=f"{percent:g}",
                expected="0 to 100",
            )
            return None
        return percent / PERCENT_DIVISOR

    def _cost(
        self,
        row: list[Any],
        index: dict[WorkbookColumn, int],
        column: WorkbookColumn,
        row_number: int,
    ) -> float | None:
        amount = self._number(row, index, column, row_number)
        if amount is None:
            if not self._is_blank(row, index, column):
                return None                       # NOT_A_NUMBER already raised
            if column in REQUIRED_COLUMNS:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.MISSING_VALUE,
                    f"{column.value} is blank.",
                    ref=self._cell(index.get(column, 0), row_number, column.value),
                )
                return None
            return 0.0
        if amount < 0:
            self._add(
                FindingSeverity.ERROR,
                FindingCode.NEGATIVE_COST,
                f"{column.value} is negative.",
                ref=self._cell(index.get(column, 0), row_number, column.value),
                supplied=f"{amount:g}",
            )
            return None
        return amount

    def _source(
        self, row: list[Any], index: dict[WorkbookColumn, int], row_number: int
    ) -> str:
        stated = self._raw(row, index, WorkbookColumn.SOURCE)
        text = "" if stated is None else str(stated).strip()
        return text or f"workbook row {row_number}"

    def _tier(
        self, row: list[Any], index: dict[WorkbookColumn, int], row_number: int
    ) -> str:
        stated = self._raw(row, index, WorkbookColumn.TIER)
        if stated is None or not str(stated).strip():
            return DEFAULT_IMPORT_TIER
        tier = str(stated).strip().upper()
        if tier not in VALID_TIERS:
            self._add(
                FindingSeverity.WARNING,
                FindingCode.UNKNOWN_TIER,
                f"{tier!r} is not a confidence tier; this row was recorded as "
                f"tier {DEFAULT_IMPORT_TIER}.",
                ref=self._cell(index.get(WorkbookColumn.TIER, 0), row_number,
                               WorkbookColumn.TIER.value),
                supplied=tier,
                expected="A, B, C or D",
            )
            return DEFAULT_IMPORT_TIER
        return tier

    # --------------------------------------------------------------- shares

    def _check_shares(self, comparators: list[ImportedComparator]) -> dict[str, float]:
        by_market: dict[str, float] = defaultdict(float)
        for comparator in comparators:
            by_market[comparator.country_code] += comparator.market_share

        for market, total in sorted(by_market.items()):
            drift = abs(total - 1.0)
            if drift <= SHARE_TOLERANCE:
                if drift > 0:
                    self._add(
                        FindingSeverity.WARNING,
                        FindingCode.SHARES_NORMALISED,
                        f"{market} shares total {total * PERCENT_DIVISOR:.1f}% and were "
                        f"normalised to 100%.",
                        supplied=f"{total * PERCENT_DIVISOR:.1f}%",
                    )
            else:
                self._add(
                    FindingSeverity.ERROR,
                    FindingCode.SHARES_DO_NOT_SUM,
                    f"{market} market shares total "
                    f"{total * PERCENT_DIVISOR:.1f}%. They must total 100% — "
                    f"{abs(1.0 - total) * PERCENT_DIVISOR:.1f}% of patients are "
                    f"{'unaccounted for' if total < 1 else 'counted twice'}.",
                    supplied=f"{total * PERCENT_DIVISOR:.1f}%",
                    expected="100%",
                )
        return dict(by_market)
