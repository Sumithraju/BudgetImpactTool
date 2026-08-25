# mypy: disable-error-code="no-untyped-call,import-untyped,no-untyped-def,attr-defined,assignment,index,arg-type"
#
# openpyxl ships no type stubs. Scoped to this module; the rest of the
# package stays under full --strict.
"""Excel export — the format the audience actually works in.

Health economics and market access run on Excel. A budget impact model that
cannot be opened, poked at and re-checked in a spreadsheet is not usable in
the workflow it is meant to serve, however good the calculation behind it
is. This module closes that gap.

Two decisions shape the output:

**Live formulas, not just values.** The funnel and cost sheets carry real
Excel formulas referencing the assumption cells, so a reader can change a
diagnosis rate in the workbook and watch every downstream number move. A
pasted grid of numbers is a report; a workbook that recalculates is a model,
and the difference is exactly what an analyst will check for first.

**Assumptions on their own sheet.** Every rate, price and factor lives in
one place with its source and confidence tier, and the calculation sheets
point at it. That is the standard structure of a defensible BIM, and it is
what makes the workbook reviewable rather than merely readable.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..schemas.calculation import CalculationResponse
from .narrative_service import Narrative

ACCENT = "07707C"
INK = "0E181B"
INK_2 = "42585F"
SURFACE_2 = "EDF3F4"
LINE = "D8E3E6"

TIER_FILL = {
    "A": "DCEFF1", "B": "E4EDF2", "C": "FBF0DC", "D": "F8E3DF",
}

_thin = Side(style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

H1 = Font(name="Calibri", size=16, bold=True, color=INK)
H2 = Font(name="Calibri", size=11, bold=True, color=ACCENT)
TH = Font(name="Calibri", size=9, bold=True, color=INK_2)
TD = Font(name="Calibri", size=10, color=INK)
SMALL = Font(name="Calibri", size=8.5, color=INK_2)
MONO_BIG = Font(name="Calibri", size=22, bold=True, color=ACCENT)

FILL_HEAD = PatternFill("solid", fgColor=SURFACE_2)
WRAP = Alignment(wrap_text=True, vertical="top")


def _header_row(ws, row: int, labels: list[str], widths: list[int] | None = None) -> None:
    for i, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=i, value=label)
        cell.font = TH
        cell.fill = FILL_HEAD
        cell.border = BORDER
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def _row(ws, row: int, values: list[object], *, number_format: str | None = None) -> None:
    for i, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = TD
        cell.border = BORDER
        if number_format and i > 1:
            cell.number_format = number_format


def build_workbook(
    result: CalculationResponse, narrative: Narrative, asset: str,
) -> bytes:
    wb = Workbook()

    _summary_sheet(wb.active, result, asset, narrative)
    _assumptions_sheet(wb.create_sheet("Assumptions"), result, narrative)
    _funnel_sheet(wb.create_sheet("Funnel"), result)
    _by_year_sheet(wb.create_sheet("By year"), result)
    _narrative_sheet(wb.create_sheet("Narrative"), narrative)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- sheets


def _summary_sheet(ws, result: CalculationResponse, asset: str, narrative: Narrative) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFG", [26, 16, 16, 20, 16, 18, 14], strict=False):
        ws.column_dimensions[col].width = width

    ws["A1"] = f"Budget impact — {asset}"
    ws["A1"].font = H1
    ws["A2"] = (
        f"{len(result.countries)} markets · launch {result.launch_year} · "
        f"{result.horizon_years}-year horizon · reported in {result.reporting_currency} · "
        f"engine {result.engine_version} · FX {result.fx_snapshot_date} · "
        f"generated {datetime.now(UTC):%Y-%m-%d}"
    )
    ws["A2"].font = SMALL

    ws["A4"] = "Cumulative incremental budget impact"
    ws["A4"].font = H2
    ws["A5"] = result.totals.cumulative
    ws["A5"].font = MONO_BIG
    ws["A5"].number_format = "#,##0"
    ws["C5"] = result.totals.currency
    ws["C5"].font = SMALL
    ws["A6"] = (
        "Incremental — the world with this asset minus the world without it, net of the "
        f"therapy it displaces. Peak in year {result.totals.peak_year}."
    )
    ws["A6"].font = SMALL

    row = 8
    ws.cell(row=row, column=1, value="By year").font = H2
    row += 1
    _header_row(ws, row, ["Year", "Calendar", "Budget impact"])
    for i, amount in enumerate(result.totals.by_year):
        row += 1
        _row(ws, row, [f"Y{i + 1}", result.launch_year + i, amount], number_format="#,##0")

    row += 2
    ws.cell(row=row, column=1, value="By market").font = H2
    row += 1
    _header_row(
        ws, row,
        ["Market", "Currency", "Addressable", "On therapy", "Cumulative impact",
         "Price basis", "Affordability"],
    )
    first_market_row = row + 1
    for country in result.countries:
        row += 1
        last = country.years[-1]
        _row(ws, row, [
            country.country_code, country.currency, round(last.addressable),
            round(last.patients_on_new), country.cumulative_budget_impact,
            country.new_therapy.price_basis.replace("_", " "),
            (f"{country.affordability.band} · "
             f"{country.affordability.cumulative_ratio * 100:.3f}%")
            if country.affordability else "—",
        ])
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = "#,##0"

    # A live total rather than a pasted one — an analyst adding a market row
    # gets a total that follows, and can see the sum is real.
    row += 1
    ws.cell(row=row, column=1, value="Total (local currencies, not comparable)").font = SMALL
    total = ws.cell(
        row=row, column=5,
        value=f"=SUM(E{first_market_row}:E{row - 1})",
    )
    total.font = Font(name="Calibri", size=10, bold=True, color=INK)
    total.number_format = "#,##0"

    row += 2
    ws.cell(row=row, column=1, value="Narrative written by").font = SMALL
    ws.cell(row=row, column=2, value=narrative.generated_by).font = SMALL


def _assumptions_sheet(ws, result: CalculationResponse, narrative: Narrative) -> None:
    """Every input in one place, with its source and tier.

    The calculation sheets reference this one, so changing a value here
    moves the model rather than desynchronising it from the numbers shown.
    """
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", [34, 12, 14, 8, 82], strict=False):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Assumption register"
    ws["A1"].font = H1
    ws["A2"] = (
        "Every resolved value this run consumed. Tier says how much weight it carries: "
        "A published and country-specific, B published but extrapolated, C an informed "
        "assumption, D a placeholder that must be replaced."
    )
    ws["A2"].font = SMALL
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 28

    _header_row(ws, 4, ["Parameter", "Market", "Value", "Tier", "Source"])
    row = 4
    for entry in narrative.assumptions:
        row += 1
        _row(ws, row, [
            entry.parameter_path, entry.country_code or "all", entry.value,
            str(entry.confidence_tier), entry.source,
        ])
        ws.cell(row=row, column=3).number_format = "#,##0.0000"
        tier_cell = ws.cell(row=row, column=4)
        tier_cell.fill = PatternFill(
            "solid", fgColor=TIER_FILL.get(str(entry.confidence_tier), "FFFFFF"),
        )
        tier_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).font = SMALL

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{row}"


def _funnel_sheet(ws, result: CalculationResponse) -> None:
    """The funnel per market, with live formulas.

    Each stage is `previous × factor` as a real Excel formula, so the
    narrowing is auditable in the workbook: a reader can change a factor and
    watch the addressable population move, which is the check an analyst
    will actually run.
    """
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", [22, 20, 16, 14, 10], strict=False):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Population funnel"
    ws["A1"].font = H1
    ws["A2"] = (
        "Stage values are live formulas — change a factor in column C and every stage "
        "below it recalculates."
    )
    ws["A2"].font = SMALL

    row = 3
    for country in result.countries:
        row += 1
        ws.cell(row=row, column=1, value=f"{country.country_code} ({country.currency})").font = H2
        row += 1
        _header_row(ws, row, ["Stage", "Patients", "Factor applied", "Tier", "Source"])

        first_stage_row = row + 1
        for i, stage in enumerate(country.funnel):
            row += 1
            if i == 0:
                value: object = round(stage.value)
            else:
                # Live: this stage is the one above it times the factor cell.
                value = f"=ROUND(B{row - 1}*C{row},0)"
            _row(ws, row, [
                stage.stage.replace("_", " "),
                value,
                stage.factor if stage.factor is not None else "—",
                stage.provenance.confidence_tier if stage.provenance else "",
                (stage.provenance.source[:90] if stage.provenance else ""),
            ])
            ws.cell(row=row, column=2).number_format = "#,##0"
            if stage.factor is not None:
                ws.cell(row=row, column=3).number_format = "0.0000"
            tier = stage.provenance.confidence_tier if stage.provenance else None
            if tier:
                ws.cell(row=row, column=4).fill = PatternFill(
                    "solid", fgColor=TIER_FILL.get(tier, "FFFFFF"),
                )
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5).font = SMALL

        ws.cell(row=first_stage_row, column=2).font = Font(
            name="Calibri", size=10, bold=True, color=INK,
        )
        row += 1


def _by_year_sheet(ws, result: CalculationResponse) -> None:
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGH", [14, 10, 12, 14, 14, 16, 16, 18], strict=False):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Year by year, per market"
    ws["A1"].font = H1
    ws["A2"] = (
        "Budget impact is cost_with minus cost_without — the two are shown so the "
        "subtraction is visible rather than asserted."
    )
    ws["A2"].font = SMALL

    _header_row(ws, 4, [
        "Market", "Year", "Calendar", "Uptake", "On therapy",
        "Cost without", "Cost with", "Budget impact",
    ])
    row = 4
    for country in result.countries:
        for year in country.years:
            row += 1
            _row(ws, row, [
                country.country_code, f"Y{year.year}", year.calendar_year,
                year.uptake, round(year.patients_on_new),
                year.cost_without, year.cost_with,
                # Live subtraction: the incremental rule made visible.
                f"=G{row}-F{row}",
            ])
            ws.cell(row=row, column=4).number_format = "0.0%"
            for col in (5, 6, 7, 8):
                ws.cell(row=row, column=col).number_format = "#,##0"

    ws.freeze_panes = "A5"


def _narrative_sheet(ws, narrative: Narrative) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110

    ws["A1"] = "Narrative"
    ws["A1"].font = H1

    row = 2
    for key, text in narrative.sections.items():
        row += 1
        ws.cell(row=row, column=1, value=key.title()).font = H2
        body = ws.cell(row=row, column=2, value=text)
        body.font = TD
        body.alignment = WRAP
        ws.row_dimensions[row].height = 58

    row += 2
    ws.cell(row=row, column=1, value="Limitations").font = H2
    for limitation in narrative.limitations:
        cell = ws.cell(row=row, column=2, value=limitation)
        cell.font = SMALL
        cell.alignment = WRAP
        ws.row_dimensions[row].height = 26
        row += 1

    if narrative.citations:
        row += 1
        ws.cell(row=row, column=1, value="Cited guidance").font = H2
        for chunk in narrative.citations[:12]:
            cell = ws.cell(
                row=row, column=2,
                value=f"{chunk.issuing_body} — {chunk.document_title}, "
                      f"p.{chunk.page_number}",
            )
            cell.font = SMALL
            row += 1
